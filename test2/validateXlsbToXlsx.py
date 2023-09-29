from openpyxl import load_workbook
import pyxlsb
from openpyxl import Workbook

# Define the path to the .xlsb file and the output .xlsx file
xlsb_file = 'C:\\Users\\dillibabu.nsp\\Downloads\\20230725155251122.xlsb'
xlsx_file = '20230725155251122.xlsx'

# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active

# Read the .xlsb file and write its contents to the Excel sheet
with pyxlsb.open_workbook(xlsb_file) as wb:
    with wb.get_sheet(1) as sheet_xlsb:
        for row_xlsb in sheet_xlsb.rows():
            row_data = [item.v for item in row_xlsb]
            sheet.append(row_data)

# Save the Excel workbook to an .xlsx file
workbook.save(xlsx_file)

print(f'Converted {xlsb_file} to {xlsx_file}')

# Load the main Excel file
main_workbook = load_workbook(xlsx_file)
main_sheet = main_workbook.active

# Load the result Excel file
result_workbook = load_workbook('C:\\Users\\dillibabu.nsp\\PycharmProjects\\DatasetsValidation\\test2\\output.xlsx')
result_sheet = result_workbook.active

# Create sets to store unique values from both files
main_values = set()
result_values = set()

# Iterate through the main Excel file to collect values
for row_main in main_sheet.iter_rows(values_only=True):
    for cell_value in row_main:
        if cell_value is not None:
            main_values.add(cell_value)

# Iterate through the result Excel file to collect values
for row_result in result_sheet.iter_rows(values_only=True):
    for cell_value in row_result:
        if cell_value is not None:
            result_values.add(cell_value)

# Check if each data point from main_values is present in result_values
for data_point in main_values:
    if data_point in result_values:
        print(f"Data point '{data_point}' from main.xlsx is present in output.xlsx.")
    else:
        print(f"Data point '{data_point}' from main.xlsx is not present in output.xlsx.")

print("Check for data presence complete")
