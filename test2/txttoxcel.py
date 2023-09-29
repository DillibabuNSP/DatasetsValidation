from openpyxl import Workbook

# Open the text file for reading
with open('C:\\Users\\dillibabu.nsp\\Downloads\\20230727200135302.txt', 'r') as text_file:
    # Create a new Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active

    # Read the text file line by line
    for line in text_file:
        # Split the line using the asterisk (*) delimiter
        data = line.strip().split('*')

        # Write the data to the Excel sheet
        sheet.append(data)

    # Create a new sheet for transposed and cleaned data
    transposed_sheet = workbook.create_sheet(title='Transposed Data')

    # Transpose the data manually
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Create a dictionary to group data by column name
    grouped_data = {}

    for col in range(1, num_cols + 1):
        col_name = sheet.cell(row=1, column=col).value  # Assuming the column names are in the first row
        if col_name not in grouped_data:
            grouped_data[col_name] = []

        for row in range(2, num_rows + 1):  # Start from the second row to skip column names
            cell_value = sheet.cell(row=row, column=col).value
            grouped_data[col_name].append(cell_value if cell_value is not None else '')

    # Write grouped and cleaned data to the transposed sheet
    for data in grouped_data.values():
        transposed_sheet.append(data)

    # Delete the original sheet
    workbook.remove(sheet)

    # Save the Excel workbook to a file
    workbook.save('output.xlsx')

print("Data has been successfully transposed and cleaned in output.xlsx")
